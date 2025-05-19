import os
from openpyxl import Workbook, load_workbook

class Trip:
    def __init__(self, distance, fuel_efficieny, fuel_price):
        self.distance = distance
        self.fuel_efficiency = fuel_efficieny
        self.fuel_price = fuel_price
        self.cost = self.calculate_cost()
    
    def calculate_cost(self):
        liters = (self.distance / 100) * self.fuel_efficiency
        return round(liters*self.fuel_efficiency, 2)

    def __repr__(self):
        return f"{self.distance} km, {self.fuel_efficiency} L/100km, {self.fuel_price} euros/l => {self.cost} euros"
    
    def to_list(self):
        return[self.distance, self.fuel_efficiency, self.fuel_price, self.cost]  
    

class BSTNode:
    def __init__(self, trip):
        self.trip = trip
        self.left = None
        self.right = None

class TripBST:
    def __init__(self):
        self.root = None

    def insert(self, trip):
        self.root = self.insert_recursive(self.root, trip)

    def _insert_recursive(self, node, trip):
        if not node:
            return BSTNode(trip)
        if trip.cost < node.trip.cost:
            node.left = self._insert_recursive(node.left, trip)
        else:
            node.right = self._insert_recursive(node.right, trip)
        return node
    
    def inorder_traversal(self):
        result =[]
        self.inorder(self.root, result)
        return result
    
    def _inorder(self, node, result):
        if node:
            self._inorder(node.left, result)
            result.append(node.trip)
            self._inorder(node.right, result)

    def find_trips_over(self, min_cost):
        result=[]
        self._collect_over_cost(self.root, min_cost, result)
        return result
    
    def _collect_over_cost(self, node, min_cost, result):
        if not node:
            return
        if node.trip.cost > min_cost:
            self._collect_over_cost(node.left, min_cost, result)
            result.append(node.trip)
            self._collect_over_cost(node.right, min_cost, result)
        else:
            self._collect_over_cost(node.right, min_cost, result)

def save_trips(trips, filename="trip_costs.xlsx"):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb= Workbook()
        ws = wb.active 
        ws.append(["Distance, Fuel efficiency, Fuel price, Trip Cost"])
    for trip in trips:
        ws.append(trip.to_list())

    wb.save(filename)

def load_trips_from_excel(filename="trip_costs.xlsx"):
    trips = []
    wb = load_workbook(filename)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        distance, efficiency, fuel_price, cost = row
        trip = Trip(distance, efficiency, fuel_price)
        trips.append(trip)

    return trips

def main():
    bst=TripBST()
    print("1 - new trip")
    print ("2 - load trips from excel")
    choice = input("").strip()

    if choice =="1":
        while True:
            d=float(input("Enter distance in km:"))
            e=float(input("Enter fuel efficiency in l/100km:"))
            p=float(input("enter fuel price in euros per liter:"))

            trip = Trip(d, e, p)
            print(f"Trip cost:{trip.cost} euros")
            bst.insert(trip)
            break
        all_trips = bst.inorder_traversal()
        save_trips(all_trips)
    elif choice=="2":
        trips = load_trips_from_excel()
        for trip in trips:
            bst.insert(trip)
        print(f"Loaded trips from excel")

        threshhold = float(input("enter minimum trip cost:"))
        filtered = bst.find_trips_over(threshhold)
        print(f"Trips over {threshhold} euros: ")
        if filtered:
            for t in filtered:
                print(t)

if __name__=="__main__":
    main()